from __future__ import annotations

import os
import re
from dataclasses import dataclass
from pathlib import Path

from database.academics import subjects


@dataclass(frozen=True)
class CurriculumSource:
    subject_name: str
    subject_short: str
    path: Path


@dataclass(frozen=True)
class CurriculumItem:
    subject_name: str
    subject_key: str
    subject_short: str
    source_file: str
    sheet_name: str
    row_number: int
    item_order: int
    lesson_number: str
    title: str
    item_type: str
    term_label: str
    week_label: str
    specification_points: str
    book_pages: str
    lesson_count: str
    duration_hours: str


# The SOW spreadsheets live in Downloads by default; override with SOW_DIR.
SOW_DIR = Path(os.environ.get("SOW_DIR", "") or (Path.home() / "Downloads"))

DEFAULT_CURRICULUM_SOURCES = (
    CurriculumSource("IGCSE Mathematics A", "Math", SOW_DIR / "IG Teaching Hubs MathsA SOW.xlsx"),
    CurriculumSource("English as a Second Language", "Eng", SOW_DIR / "IG Teaching Hubs ESL SOW.xlsx"),
    CurriculumSource("IGCSE Chemistry", "Chem", SOW_DIR / "IG Teaching Hubs Chemistry SOW.xlsx"),
    CurriculumSource("IGCSE Biology", "Bio", SOW_DIR / "IG Teaching Hubs Biology SOW.xlsx"),
    CurriculumSource("IGCSE Physics", "Phy", SOW_DIR / "IG Teaching Hubs Physics SOW.xlsx"),
)

_EXAM_KEYWORDS = (
    "half-term test",
    "end-of-term test",
    "end of term test",
    "end of year test",
    "mock exam",
    "mock unit test",
    "exam practice",
)

_LESSON_NUMBER_RE = re.compile(r"\bLesson\s+(\d+)\b", re.IGNORECASE)


def _clean(value: object) -> str:
    return re.sub(r"\s+", " ", str("" if value is None else value).strip())


def classify_curriculum_item(title: str) -> str:
    text = _clean(title).casefold()
    return "exam" if any(keyword in text for keyword in _EXAM_KEYWORDS) else "lesson"


def _read_rows(path: Path) -> tuple[str, list[tuple[int, dict[int, str]]]]:
    from openpyxl import load_workbook  # imported lazily; only needed when parsing

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = []
        for cells in sheet.iter_rows():
            filled = [cell for cell in cells if cell.value is not None]
            if filled:
                rows.append((filled[0].row, {cell.column: _clean(cell.value) for cell in filled}))
        return sheet.title, rows
    finally:
        workbook.close()


def _find_header(rows: list[tuple[int, dict[int, str]]]) -> tuple[int, dict[int, str]]:
    """The header row is the one that has both a 'lesson number' and 'lesson name' cell."""
    for row_number, values in rows:
        header = {column: text.casefold() for column, text in values.items()}
        if "lesson number" in header.values() and "lesson name" in header.values():
            return row_number, header
    raise ValueError("Could not find the lesson header row.")


def _column_for(header: dict[int, str], *labels: str) -> int | None:
    wanted = {label.casefold() for label in labels}
    for column, text in header.items():
        if text in wanted:
            return column
    return None


def parse_curriculum_source(source: CurriculumSource) -> list[CurriculumItem]:
    if not source.path.exists():
        raise FileNotFoundError(source.path)

    sheet_name, rows = _read_rows(source.path)
    header_row, header = _find_header(rows)

    number_col = _column_for(header, "lesson number")
    title_col = _column_for(header, "lesson name")
    if number_col is None or title_col is None:
        raise ValueError("Lesson number/title columns are required.")

    term_col = _column_for(header, "year and term")
    week_col = _column_for(header, "week")
    spec_col = _column_for(header, "specification point(s)", "assessment objectives")
    pages_col = _column_for(header, "student book pages", "student book page reference")
    count_col = _column_for(header, "no. of lesson")
    hours_col = _column_for(header, "no. of hours (0.66 = 40 mins)")

    subject_name = subjects.canonical_subject_name(source.subject_name)
    subject_key = subjects.subject_key(subject_name)
    subject_short = source.subject_short or subjects.subject_short_name(subject_name)

    def cell(values: dict[int, str], column: int | None) -> str:
        return values.get(column, "") if column else ""

    items: list[CurriculumItem] = []
    for row_number, values in rows:
        if row_number <= header_row:
            continue

        lesson_number = values.get(number_col, "")
        title = values.get(title_col, "")
        match = _LESSON_NUMBER_RE.search(lesson_number)
        if not match:
            continue

        order = int(match.group(1))
        items.append(
            CurriculumItem(
                subject_name=subject_name,
                subject_key=subject_key,
                subject_short=subject_short,
                source_file=source.path.name,
                sheet_name=sheet_name,
                row_number=row_number,
                item_order=order,
                lesson_number=f"Lesson {order}",
                title=title,
                item_type=classify_curriculum_item(title),
                term_label=cell(values, term_col),
                week_label=cell(values, week_col),
                specification_points=cell(values, spec_col),
                book_pages=cell(values, pages_col),
                lesson_count=cell(values, count_col),
                duration_hours=cell(values, hours_col),
            )
        )

    return sorted(items, key=lambda item: item.item_order)


def load_default_curricula() -> dict[str, list[CurriculumItem]]:
    curricula: dict[str, list[CurriculumItem]] = {}
    for source in DEFAULT_CURRICULUM_SOURCES:
        items = parse_curriculum_source(source)
        if items:
            curricula[items[0].subject_key] = items
    return curricula


__all__ = [
    "CurriculumItem",
    "CurriculumSource",
    "DEFAULT_CURRICULUM_SOURCES",
    "classify_curriculum_item",
    "load_default_curricula",
    "parse_curriculum_source",
]
