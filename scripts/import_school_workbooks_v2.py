from __future__ import annotations

import argparse
import hashlib
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database import connect_auth_db
from database.academics import dates, schools, subjects


SCHOOL5_GROUP_SHEETS = {
    "MMG1": ("MG1", "IGCSE Mathematics A"),
    "MMG2": ("MG2", "IGCSE Mathematics A"),
    "MAFTG1": ("AFT1", "IGCSE Mathematics A"),
    "MAFTG2": ("AFT2", "IGCSE Mathematics A"),
    "MONLINE": ("ONLINE", "IGCSE Mathematics A"),
    "ENGMG1": ("MG1", "English as a Second Language"),
    "ENGAFTG1": ("AFT1", "English as a Second Language"),
}

SEHRIYO_GROUP_RE = re.compile(r"^(?P<group>[78][A-ZА-ЯVGDБВГД]+)\s*-\s*(?P<subject>Math|CH|Ch)$", re.IGNORECASE)


@dataclass
class StudentRecord:
    school_key: str
    school_name: str
    full_name: str
    student_code: str = ""
    password_plain: str = ""
    subjects: set[str] = field(default_factory=set)
    groups: set[tuple[str, str]] = field(default_factory=set)


@dataclass
class GroupRecord:
    school_key: str
    school_name: str
    group_name: str
    subject_name: str
    source_sheet: str
    students: list[str] = field(default_factory=list)


@dataclass
class ExamRecord:
    school_key: str
    group_name: str
    subject_name: str
    student_name: str
    lesson_number: str
    exam_name: str
    attempt: str
    score: float
    exam_date: date | None
    source_sheet: str
    source_column: int


@dataclass
class WorkbookImport:
    students: dict[tuple[str, str], StudentRecord] = field(default_factory=dict)
    groups: dict[tuple[str, str, str], GroupRecord] = field(default_factory=dict)
    exams: list[ExamRecord] = field(default_factory=list)


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str("" if value is None else value).strip())


def _name_key(value: str) -> str:
    return _clean(value).casefold()


def _is_student_row_number(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return int(value) == value and value > 0
    text = _clean(value)
    return bool(re.fullmatch(r"\d+(?:\.0)?", text))


def _number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean(value).replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = dates.parse_date(value)
    return parsed


def _sheet_rows(ws) -> list[list[Any]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def _cell(rows: list[list[Any]], row_index: int, col_index: int) -> Any:
    if row_index >= len(rows):
        return None
    row = rows[row_index]
    if col_index >= len(row):
        return None
    return row[col_index]


def _canonical_subject(value: str) -> str:
    return subjects.canonical_subject_name(value)


def _subject_short(value: str) -> str:
    return subjects.subject_short_name(_canonical_subject(value))


def _subject_from_sehriyo_suffix(value: str) -> str:
    normalized = value.strip().casefold()
    if normalized == "math":
        return "IGCSE Mathematics A"
    if normalized in {"ch", "chem"}:
        return "IGCSE Chemistry"
    return value


def _school_student_prefix(school_key: str) -> str:
    return schools.student_code_prefix(school_key)


def _student_key(school_key: str, full_name: str) -> tuple[str, str]:
    return (schools.normalize_school_code(school_key), _name_key(full_name))


def _group_key(school_key: str, subject_name: str, group_name: str) -> tuple[str, str, str]:
    return (
        schools.normalize_school_code(school_key),
        subjects.subject_key(subject_name),
        _clean(group_name).casefold(),
    )


def _stable_dashboard_id(school_key: str, group_name: str, subject_name: str, full_name: str) -> int:
    text = "|".join(
        [
            schools.normalize_school_code(school_key),
            _clean(group_name).casefold(),
            subjects.subject_key(subject_name),
            _name_key(full_name),
        ]
    )
    digest = hashlib.sha1(text.encode("utf-8")).hexdigest()
    return 9_000_000_000 + int(digest[:8], 16)


def _register_student(
    result: WorkbookImport,
    *,
    school_key: str,
    school_name: str,
    full_name: str,
    subject_name: str = "",
    group_name: str = "",
    student_code: str = "",
    password_plain: str = "",
) -> StudentRecord:
    key = _student_key(school_key, full_name)
    record = result.students.get(key)
    if record is None:
        record = StudentRecord(
            school_key=schools.normalize_school_code(school_key),
            school_name=school_name,
            full_name=_clean(full_name),
            student_code=_clean(student_code).upper(),
            password_plain=_clean(password_plain),
        )
        result.students[key] = record
    if student_code and not record.student_code:
        record.student_code = _clean(student_code).upper()
    if password_plain and not record.password_plain:
        record.password_plain = _clean(password_plain)
    if subject_name:
        canonical_subject = _canonical_subject(subject_name)
        record.subjects.add(canonical_subject)
        if group_name:
            record.groups.add((canonical_subject, _clean(group_name)))
    return record


def _parse_auth_sheet(result: WorkbookImport, path: Path, *, school_key: str, school_name: str) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Auth" not in wb.sheetnames:
            return
        ws = wb["Auth"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            full_name = _clean(row[1] if len(row) > 1 else "")
            student_code = _clean(row[2] if len(row) > 2 else "")
            password = _clean(row[3] if len(row) > 3 else "")
            subject_text = _clean(row[4] if len(row) > 4 else "")
            if not full_name or not student_code:
                continue
            record = _register_student(
                result,
                school_key=school_key,
                school_name=school_name,
                full_name=full_name,
                student_code=student_code,
                password_plain=password or student_code,
            )
            for part in subjects.split_subjects(subject_text):
                record.subjects.add(_canonical_subject(part))
    finally:
        wb.close()


def _parse_group_sheet(
    result: WorkbookImport,
    path: Path,
    *,
    school_key: str,
    school_name: str,
    sheet_name: str,
    group_name: str,
    subject_name: str,
) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = _sheet_rows(ws)
    finally:
        wb.close()

    canonical_subject = _canonical_subject(subject_name)
    clean_group = _clean(group_name)
    group = result.groups.setdefault(
        _group_key(school_key, canonical_subject, clean_group),
        GroupRecord(
            school_key=schools.normalize_school_code(school_key),
            school_name=school_name,
            group_name=clean_group,
            subject_name=canonical_subject,
            source_sheet=sheet_name,
        ),
    )

    date_row = rows[0] if rows else []
    lesson_number_row = rows[1] if len(rows) > 1 else []
    lesson_name_row = rows[2] if len(rows) > 2 else []
    aap_col_index = None
    for header_row in (date_row, lesson_number_row, lesson_name_row):
        for col_index, value in enumerate(header_row):
            if _clean(value).casefold() == "aap":
                aap_col_index = col_index
                break
        if aap_col_index is not None:
            break
    exam_col_limit = aap_col_index if aap_col_index is not None else len(date_row)

    for row_index in range(3, len(rows)):
        row = rows[row_index]
        if len(row) < 2 or not _is_student_row_number(row[0]):
            continue
        full_name = _clean(row[1])
        if not full_name:
            continue
        group.students.append(full_name)
        _register_student(
            result,
            school_key=school_key,
            school_name=school_name,
            full_name=full_name,
            subject_name=canonical_subject,
            group_name=clean_group,
        )

        last_lesson_number = ""
        last_exam_name = ""
        last_date: date | None = None
        for col_index in range(2, min(len(row), exam_col_limit)):
            lesson_number = _clean(lesson_number_row[col_index] if col_index < len(lesson_number_row) else "")
            exam_name = _clean(lesson_name_row[col_index] if col_index < len(lesson_name_row) else "")
            exam_date = _date(date_row[col_index] if col_index < len(date_row) else None)
            if lesson_number:
                last_lesson_number = lesson_number
            if exam_name:
                last_exam_name = exam_name
            if exam_date:
                last_date = exam_date
            score = _number(row[col_index])
            if score is None:
                continue
            if score < 0 or score > 9:
                continue
            result.exams.append(
                ExamRecord(
                    school_key=schools.normalize_school_code(school_key),
                    group_name=clean_group,
                    subject_name=canonical_subject,
                    student_name=full_name,
                    lesson_number=last_lesson_number or f"Column {col_index + 1}",
                    exam_name=last_exam_name or last_lesson_number or f"Column {col_index + 1}",
                    attempt=f"{sheet_name}!{col_index + 1}",
                    score=score,
                    exam_date=last_date,
                    source_sheet=sheet_name,
                    source_column=col_index + 1,
                )
            )


def parse_workbooks(school5_path: Path, sehriyo_path: Path) -> WorkbookImport:
    result = WorkbookImport()
    _parse_auth_sheet(result, school5_path, school_key="school5", school_name="School 5")

    for sheet_name, (group_name, subject_name) in SCHOOL5_GROUP_SHEETS.items():
        _parse_group_sheet(
            result,
            school5_path,
            school_key="school5",
            school_name="School 5",
            sheet_name=sheet_name,
            group_name=group_name,
            subject_name=subject_name,
        )

    wb = load_workbook(sehriyo_path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
    finally:
        wb.close()
    for sheet_name in sheet_names:
        match = SEHRIYO_GROUP_RE.match(sheet_name.strip())
        if not match:
            continue
        _parse_group_sheet(
            result,
            sehriyo_path,
            school_key="sehriyo",
            school_name="Sehriyo",
            sheet_name=sheet_name,
            group_name=match.group("group").replace(" ", ""),
            subject_name=_subject_from_sehriyo_suffix(match.group("subject")),
        )

    return result


def _row_value(row, key: str, default=None):
    return row[key] if row and key in row else default


def _fetch_id(conn, sql: str, params: tuple[Any, ...]) -> int:
    row = conn.execute(sql, params).fetchone()
    if not row:
        raise RuntimeError(f"Could not fetch id for SQL: {sql}")
    return int(row["id"])


def _ensure_school(conn, school_key: str, school_name: str) -> int:
    conn.execute(
        """
        INSERT INTO msi_v2.schools (school_key, school_name, status, created_at, updated_at)
        VALUES (%s, %s, 'active', now(), now())
        ON CONFLICT ((lower(school_key))) DO UPDATE SET
            school_name = EXCLUDED.school_name,
            status = 'active',
            updated_at = now()
        """,
        (school_key, school_name),
    )
    return _fetch_id(conn, "SELECT id FROM msi_v2.schools WHERE lower(school_key) = lower(%s)", (school_key,))


def _ensure_subject(conn, subject_name: str) -> int:
    canonical = _canonical_subject(subject_name)
    key = subjects.subject_key(canonical)
    conn.execute(
        """
        INSERT INTO msi_v2.subjects (subject_key, subject_name, subject_short, status, created_at, updated_at)
        VALUES (%s, %s, %s, 'active', now(), now())
        ON CONFLICT ((lower(subject_key))) DO UPDATE SET
            subject_name = EXCLUDED.subject_name,
            subject_short = EXCLUDED.subject_short,
            status = 'active',
            updated_at = now()
        """,
        (key, canonical, _subject_short(canonical)),
    )
    return _fetch_id(conn, "SELECT id FROM msi_v2.subjects WHERE lower(subject_key) = lower(%s)", (key,))


def _ensure_program(conn, subject_id: int, subject_name: str) -> int:
    conn.execute(
        """
        INSERT INTO msi_v2.subject_programs (
            subject_id, academic_year, program_name, source_file, status, created_at, updated_at
        )
        VALUES (%s, '2025-26', %s, 'official SOW', 'active', now(), now())
        ON CONFLICT (subject_id, academic_year) DO UPDATE SET
            program_name = EXCLUDED.program_name,
            status = 'active',
            updated_at = now()
        """,
        (subject_id, f"{subject_name} 2025-26"),
    )
    return _fetch_id(
        conn,
        "SELECT id FROM msi_v2.subject_programs WHERE subject_id = %s AND academic_year = '2025-26'",
        (subject_id,),
    )


def _next_legacy_student_row_id(conn) -> int:
    row = conn.execute(
        "SELECT COALESCE(MAX(legacy_student_row_id), 0) + 1 AS next_id FROM msi_v2.students"
    ).fetchone()
    return int(row["next_id"] or 1)


def _next_student_code(conn, prefix: str) -> str:
    rows = conn.execute(
        "SELECT student_code FROM msi_v2.students WHERE upper(student_code) LIKE %s",
        (f"{prefix.upper()}%",),
    ).fetchall()
    max_num = 0
    for row in rows:
        text = _clean(row["student_code"]).upper()
        suffix = text[len(prefix):]
        if text.startswith(prefix.upper()) and suffix.isdigit():
            max_num = max(max_num, int(suffix))
    return f"{prefix.upper()}{max_num + 1:05d}"


def _find_student(conn, school_id: int, record: StudentRecord):
    if record.student_code:
        row = conn.execute(
            "SELECT id, student_code, legacy_student_row_id FROM msi_v2.students WHERE upper(student_code) = upper(%s)",
            (record.student_code,),
        ).fetchone()
        if row:
            return row
    return conn.execute(
        """
        SELECT id, student_code, legacy_student_row_id
        FROM msi_v2.students
        WHERE school_id = %s AND lower(full_name) = lower(%s)
        ORDER BY id ASC
        LIMIT 1
        """,
        (school_id, record.full_name),
    ).fetchone()


def _ensure_student(conn, school_id: int, record: StudentRecord, counters: dict[str, int]) -> int:
    existing = _find_student(conn, school_id, record)
    if existing:
        student_id = int(existing["id"])
        code = record.student_code or _clean(existing["student_code"])
        password = record.password_plain or code
        conn.execute(
            """
            UPDATE msi_v2.students
            SET
                student_code = %s,
                full_name = %s,
                school_id = %s,
                password_plain = CASE WHEN COALESCE(password_plain, '') = '' THEN %s ELSE password_plain END,
                status = 'active',
                updated_at = now(),
                legacy_student_row_id = COALESCE(legacy_student_row_id, %s)
            WHERE id = %s
            """,
            (code, record.full_name, school_id, password, _next_legacy_student_row_id(conn), student_id),
        )
        counters["students_updated"] += 1
    else:
        if not record.student_code:
            record.student_code = _next_student_code(conn, _school_student_prefix(record.school_key))
        if not record.password_plain:
            record.password_plain = record.student_code
        legacy_id = _next_legacy_student_row_id(conn)
        row = conn.execute(
            """
            INSERT INTO msi_v2.students (
                student_code, full_name, school_id, password_plain, status,
                legacy_student_row_id, created_at, updated_at
            )
            VALUES (%s, %s, %s, %s, 'active', %s, now(), now())
            RETURNING id
            """,
            (record.student_code, record.full_name, school_id, record.password_plain, legacy_id),
        ).fetchone()
        student_id = int(row["id"])
        counters["students_created"] += 1

    conn.execute(
        """
        INSERT INTO msi_v2.student_auth (student_id, password_hash, must_change_password, updated_at)
        VALUES (%s, %s, true, now())
        ON CONFLICT (student_id) DO NOTHING
        """,
        (student_id, generate_password_hash(record.password_plain or record.student_code)),
    )
    return student_id


def _ensure_group(conn, school_id: int, program_id: int, group: GroupRecord) -> int:
    row = conn.execute(
        """
        SELECT id
        FROM msi_v2.groups
        WHERE school_id = %s AND program_id = %s AND lower(group_name) = lower(%s)
        ORDER BY id ASC
        LIMIT 1
        """,
        (school_id, program_id, group.group_name),
    ).fetchone()
    if row:
        group_id = int(row["id"])
        conn.execute(
            """
            UPDATE msi_v2.groups
            SET group_name = %s, group_code = %s, status = 'active', updated_at = now()
            WHERE id = %s
            """,
            (group.group_name, f"{group.group_name}-{_subject_short(group.subject_name)}", group_id),
        )
        return group_id
    row = conn.execute(
        """
        INSERT INTO msi_v2.groups (
            school_id, program_id, group_name, group_code, status, created_at, updated_at
        )
        VALUES (%s, %s, %s, %s, 'active', now(), now())
        RETURNING id
        """,
        (school_id, program_id, group.group_name, f"{group.group_name}-{_subject_short(group.subject_name)}"),
    ).fetchone()
    return int(row["id"])


def _ensure_program_item_for_exam(conn, program_id: int, exam: ExamRecord) -> int | None:
    match = re.search(r"(\d+)", exam.lesson_number or "")
    if not match:
        return None
    item_order = int(match.group(1))
    row = conn.execute(
        """
        SELECT id
        FROM msi_v2.subject_program_items
        WHERE program_id = %s AND item_order = %s
        LIMIT 1
        """,
        (program_id, item_order),
    ).fetchone()
    if row:
        return int(row["id"])
    row = conn.execute(
        """
        INSERT INTO msi_v2.subject_program_items (
            program_id, item_order, lesson_number, item_type, title,
            source_file, sheet_name, source_row, created_at, updated_at
        )
        VALUES (%s, %s, %s, 'exam', %s, 'school workbook import', %s, 0, now(), now())
        ON CONFLICT (program_id, item_order) DO UPDATE SET
            title = EXCLUDED.title,
            item_type = 'exam',
            updated_at = now()
        RETURNING id
        """,
        (program_id, item_order, f"Lesson {item_order}", exam.exam_name, exam.source_sheet),
    ).fetchone()
    return int(row["id"])


def _ensure_exam_result(conn, group_id: int, student_id: int, program_item_id: int | None, exam: ExamRecord, counters: dict[str, int]) -> None:
    existing = conn.execute(
        """
        SELECT id
        FROM msi_v2.exam_results
        WHERE group_id = %s
          AND student_id = %s
          AND COALESCE(program_item_id, 0) = COALESCE(%s, 0)
          AND exam_name = %s
          AND attempt = %s
        ORDER BY id ASC
        LIMIT 1
        """,
        (group_id, student_id, program_item_id, exam.exam_name, exam.attempt),
    ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE msi_v2.exam_results
            SET score = %s, score_scale = 9, updated_at = now()
            WHERE id = %s
            """,
            (exam.score, int(existing["id"])),
        )
        counters["exam_results_updated"] += 1
        return
    conn.execute(
        """
        INSERT INTO msi_v2.exam_results (
            group_id, program_item_id, student_id, exam_name, attempt, score, score_scale, created_at, updated_at
        )
        VALUES (%s, %s, %s, %s, %s, %s, 9, now(), now())
        """,
        (group_id, program_item_id, student_id, exam.exam_name, exam.attempt, exam.score),
    )
    counters["exam_results_created"] += 1


def preview_import(import_data: WorkbookImport) -> dict[str, Any]:
    by_school: dict[str, int] = {}
    by_group: dict[str, int] = {}
    by_subject: dict[str, int] = {}
    for record in import_data.students.values():
        by_school[record.school_key] = by_school.get(record.school_key, 0) + 1
    for group in import_data.groups.values():
        label = f"{group.school_key}:{group.group_name}:{_subject_short(group.subject_name)}"
        by_group[label] = len(set(_name_key(name) for name in group.students))
    for exam in import_data.exams:
        by_subject[exam.subject_name] = by_subject.get(exam.subject_name, 0) + 1
    return {
        "students": len(import_data.students),
        "groups": len(import_data.groups),
        "exam_scores": len(import_data.exams),
        "students_by_school": by_school,
        "groups": by_group,
        "exam_scores_by_subject": by_subject,
    }


def apply_import(import_data: WorkbookImport) -> dict[str, int]:
    counters = {
        "students_created": 0,
        "students_updated": 0,
        "students_inactivated": 0,
        "groups_linked": 0,
        "memberships_inactivated": 0,
        "groups_created_or_updated": 0,
        "exam_results_created": 0,
        "exam_results_updated": 0,
    }
    with connect_auth_db() as conn:
        school_ids = {
            key: _ensure_school(conn, key, schools.school_display_name(key))
            for key in sorted({record.school_key for record in import_data.students.values()})
        }
        subject_ids: dict[str, int] = {}
        program_ids: dict[str, int] = {}
        for subject_name in sorted({group.subject_name for group in import_data.groups.values()}):
            subject_id = _ensure_subject(conn, subject_name)
            subject_ids[subjects.subject_key(subject_name)] = subject_id
            program_ids[subjects.subject_key(subject_name)] = _ensure_program(conn, subject_id, subject_name)

        student_ids: dict[tuple[str, str], int] = {}
        for key, record in sorted(import_data.students.items(), key=lambda item: (item[1].school_key, item[1].full_name)):
            student_ids[key] = _ensure_student(conn, school_ids[record.school_key], record, counters)
        for school_key, school_id in school_ids.items():
            active_student_ids = [
                student_id
                for key, student_id in student_ids.items()
                if key[0] == school_key
            ]
            if active_student_ids:
                result = conn.execute(
                    """
                    UPDATE msi_v2.students
                    SET status = 'inactive', updated_at = now()
                    WHERE school_id = %s
                      AND status = 'active'
                      AND NOT (id = ANY(%s::bigint[]))
                    """,
                    (school_id, active_student_ids),
                )
            else:
                result = conn.execute(
                    """
                    UPDATE msi_v2.students
                    SET status = 'inactive', updated_at = now()
                    WHERE school_id = %s AND status = 'active'
                    """,
                    (school_id,),
                )
            counters["students_inactivated"] += max(int(result.rowcount or 0), 0)

        group_ids: dict[tuple[str, str, str], int] = {}
        for key, group in sorted(import_data.groups.items(), key=lambda item: item[0]):
            program_id = program_ids[subjects.subject_key(group.subject_name)]
            group_id = _ensure_group(conn, school_ids[group.school_key], program_id, group)
            group_ids[key] = group_id
            counters["groups_created_or_updated"] += 1
            active_student_ids = [
                student_ids[_student_key(group.school_key, full_name)]
                for full_name in sorted(set(group.students), key=_name_key)
            ]
            if active_student_ids:
                result = conn.execute(
                    """
                    UPDATE msi_v2.group_students
                    SET enrollment_status = 'inactive', left_at = COALESCE(left_at, now())
                    WHERE group_id = %s
                      AND enrollment_status = 'active'
                      AND NOT (student_id = ANY(%s::bigint[]))
                    """,
                    (group_id, active_student_ids),
                )
            else:
                result = conn.execute(
                    """
                    UPDATE msi_v2.group_students
                    SET enrollment_status = 'inactive', left_at = COALESCE(left_at, now())
                    WHERE group_id = %s AND enrollment_status = 'active'
                    """,
                    (group_id,),
                )
            counters["memberships_inactivated"] += max(int(result.rowcount or 0), 0)
            for student_id, full_name in zip(active_student_ids, sorted(set(group.students), key=_name_key)):
                dashboard_id = _stable_dashboard_id(group.school_key, group.group_name, group.subject_name, full_name)
                conn.execute(
                    """
                    INSERT INTO msi_v2.group_students (
                        group_id, student_id, enrollment_status, joined_at, legacy_public_dashboard_id
                    )
                    VALUES (%s, %s, 'active', now(), %s)
                    ON CONFLICT (group_id, student_id) DO UPDATE SET
                        enrollment_status = 'active',
                        left_at = NULL,
                        legacy_public_dashboard_id = COALESCE(
                            msi_v2.group_students.legacy_public_dashboard_id,
                            EXCLUDED.legacy_public_dashboard_id
                        )
                    """,
                    (group_id, student_id, dashboard_id),
                )
                counters["groups_linked"] += 1

        program_item_cache: dict[tuple[str, str, int], int | None] = {}
        for exam in import_data.exams:
            group_id = group_ids[_group_key(exam.school_key, exam.subject_name, exam.group_name)]
            student_id = student_ids[_student_key(exam.school_key, exam.student_name)]
            match = re.search(r"(\d+)", exam.lesson_number or "")
            cache_key = (subjects.subject_key(exam.subject_name), exam.lesson_number, int(match.group(1)) if match else 0)
            if cache_key not in program_item_cache:
                program_item_cache[cache_key] = _ensure_program_item_for_exam(
                    conn,
                    program_ids[subjects.subject_key(exam.subject_name)],
                    exam,
                )
            _ensure_exam_result(conn, group_id, student_id, program_item_cache[cache_key], exam, counters)

        conn.commit()
    return counters


def main() -> int:
    parser = argparse.ArgumentParser(description="Import School 5 and Sehriyo student workbooks into msi_v2.")
    parser.add_argument("--school5", default="/Users/apple/Downloads/School_5_MSI.xlsx")
    parser.add_argument("--sehriyo", default="/Users/apple/Downloads/Sehriyo_MSI (2).xlsx")
    parser.add_argument("--apply", action="store_true", help="Write changes to PostgreSQL. Without this, only preview.")
    args = parser.parse_args()

    school5_path = Path(args.school5).expanduser()
    sehriyo_path = Path(args.sehriyo).expanduser()
    for path in (school5_path, sehriyo_path):
        if not path.exists():
            raise FileNotFoundError(path)

    import_data = parse_workbooks(school5_path, sehriyo_path)
    preview = preview_import(import_data)
    print("Import preview")
    print("==============")
    print(f"students: {preview['students']}")
    print(f"groups: {len(preview['groups'])}")
    print(f"exam_scores: {preview['exam_scores']}")
    print("students_by_school:", preview["students_by_school"])
    print("exam_scores_by_subject:", preview["exam_scores_by_subject"])
    print("groups:")
    for group_name, total in sorted(preview["groups"].items()):
        print(f"  {group_name}: {total}")

    if not args.apply:
        print("\nDry run only. Re-run with --apply to write changes.")
        return 0

    counters = apply_import(import_data)
    print("\nApplied")
    print("=======")
    for key, value in counters.items():
        print(f"{key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
