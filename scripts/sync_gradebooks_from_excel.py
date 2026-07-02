#!/usr/bin/env python3
"""Sync MSI gradebooks from the legacy Excel workbooks.

The source sheets are wide gradebooks:
- rows 1-3 contain dates, lesson labels, and topics
- rows 4..N contain active student values
- a later "Lesson AAP" row marks the end of active students
- the first "AAP" column separates exam columns from the conducted lesson grid

The script is intentionally conservative with identities: it syncs values for
students already enrolled in the matching group and reports unmatched source
students instead of creating accounts implicitly.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from database import queries
from database.academics import canonical


DEFAULT_SOURCES = [
    {
        "path": Path("/Users/apple/Downloads/School_5_MSI.xlsx"),
        "school_key": "school5",
        "sheets": {
            "MMG1": ("MG1", "IGCSE Mathematics A"),
            "MMG2": ("MG2", "IGCSE Mathematics A"),
            "MAFTG1": ("AFT1", "IGCSE Mathematics A"),
            "MAFTG2": ("AFT2", "IGCSE Mathematics A"),
            "ENGMG1": ("MG1", "English as a Second Language"),
            "ENGAFTG1": ("AFT1", "English as a Second Language"),
            "MONLINE": ("ONLINE", "IGCSE Mathematics A"),
        },
    },
    {
        "path": Path("/Users/apple/Downloads/Sehriyo_MSI (2).xlsx"),
        "school_key": "sehriyo",
        "sheets": {
            "7A-Math": ("7A", "IGCSE Mathematics A"),
            "7B-Math": ("7B", "IGCSE Mathematics A"),
            "7V-Math": ("7V", "IGCSE Mathematics A"),
            "7G-Math": ("7G", "IGCSE Mathematics A"),
            "7D - Math": ("7D", "IGCSE Mathematics A"),
            "8A-Math": ("8A", "IGCSE Mathematics A"),
            "8B-Math": ("8B", "IGCSE Mathematics A"),
            "8G-Math": ("8G", "IGCSE Mathematics A"),
            "8D-Math": ("8D", "IGCSE Mathematics A"),
            "8A-CH": ("8A", "IGCSE Chemistry"),
            "8B-CH": ("8B", "IGCSE Chemistry"),
            "8G-CH": ("8G", "IGCSE Chemistry"),
            "8D-Ch": ("8D", "IGCSE Chemistry"),
        },
    },
]


@dataclass(frozen=True)
class SourceStudent:
    row: int
    name: str
    key: str


@dataclass(frozen=True)
class SourceSession:
    column: int
    order: int
    label: str
    topic: str
    kind: str
    lesson_order: int
    session_date: date | None
    attendance_col: int
    homework_col: int | None


@dataclass(frozen=True)
class SourceExam:
    column: int
    order: int
    label: str
    topic: str
    attempt: str
    lesson_order: int
    exam_date: date | None

    @property
    def session_date(self) -> date | None:
        return self.exam_date


def compact_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").replace("\t", " ").strip().split())


def normalize_key(value: Any) -> str:
    return canonical.normalize_text(compact_text(value))


def parse_lesson_order(value: Any) -> int:
    match = re.search(r"(\d+)", compact_text(value))
    return int(match.group(1)) if match else 0


def parse_source_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = compact_text(value)
    if not text:
        return None
    parsed = canonical.parse_date(text)
    if parsed:
        return parsed
    # Excel occasionally has values such as "04 -02-2026"; canonical parsing
    # handles most spacing, but keep one extra compact fallback here.
    compact = re.sub(r"\s+", "", text)
    if compact != text:
        return canonical.parse_date(compact)
    return None


def parse_score(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        score = float(value)
    except (TypeError, ValueError):
        return None
    if not 0 <= score <= 9:
        return None
    return round(score, 1)


def normalize_attendance(value: Any) -> str:
    text = normalize_key(value).replace(" ", "")
    if not text:
        return ""
    if text == "p":
        return "present"
    if text == "a":
        return "absent"
    if text in {"a(i)", "ai", "aі", "l", "late", "j", "justified"}:
        return "justified"
    return ""


def find_aap_column(ws) -> int:
    for column in range(1, ws.max_column + 1):
        if normalize_key(ws.cell(1, column).value) == "aap":
            return column
    for column in range(4, ws.max_column + 1):
        if normalize_key(ws.cell(2, column).value) == "lesson 1":
            return column - 1
    raise ValueError(f"{ws.title}: cannot find AAP marker column")


def find_lesson_aap_row(ws) -> int:
    for row in range(1, ws.max_row + 1):
        if normalize_key(ws.cell(row, 1).value) in {"lesson aap", "lesson average aap"}:
            return row
    student_rows = [
        row
        for row in range(4, ws.max_row + 1)
        if compact_text(ws.cell(row, 2).value) and parse_score(ws.cell(row, 1).value) is not None
    ]
    if student_rows:
        return max(student_rows) + 1
    raise ValueError(f"{ws.title}: cannot find Lesson AAP row")


def source_students(ws, lesson_aap_row: int) -> list[SourceStudent]:
    students: list[SourceStudent] = []
    for row in range(4, lesson_aap_row):
        number = ws.cell(row, 1).value
        name = compact_text(ws.cell(row, 2).value)
        if not name:
            continue
        try:
            float(number)
        except (TypeError, ValueError):
            continue
        students.append(SourceStudent(row=row, name=name, key=normalize_key(name)))
    return students


def source_kind(label: str) -> str:
    key = normalize_key(label)
    if "cancelled" in key or "canceled" in key:
        return "cancelled"
    if "practice" in key:
        return "practice"
    if re.search(r"\blesson\s+\d+\b", key):
        return "lesson"
    return ""


def is_exact_lesson_label(label: str) -> bool:
    return re.fullmatch(r"lesson\s+\d+", normalize_key(label)) is not None


def source_sessions(ws, aap_col: int) -> list[SourceSession]:
    sessions: list[SourceSession] = []
    order = 0
    for column in range(aap_col + 1, ws.max_column + 1):
        label = compact_text(ws.cell(2, column).value)
        kind = source_kind(label)
        if not kind:
            continue
        order += 10
        next_label = compact_text(ws.cell(2, column + 1).value) if column + 1 <= ws.max_column else ""
        homework_col = None
        if kind == "lesson" and not next_label:
            homework_col = column + 1
        sessions.append(
            SourceSession(
                column=column,
                order=order,
                label=label,
                topic=compact_text(ws.cell(3, column).value),
                kind=kind,
                lesson_order=parse_lesson_order(label),
                session_date=parse_source_date(ws.cell(1, column).value),
                attendance_col=column,
                homework_col=homework_col,
            )
        )
    return sessions


def attempt_labels(ws, student_end_row: int, lesson_aap_row: int) -> dict[int, str]:
    labels: dict[int, str] = {}
    for row in range(student_end_row + 1, lesson_aap_row):
        for column in range(1, ws.max_column + 1):
            text = normalize_key(ws.cell(row, column).value)
            if text in {"1st", "first", "1"}:
                labels[column] = "1st"
            elif text in {"2nd", "second", "2"}:
                labels[column] = "2nd"
    return labels


def source_exams(ws, aap_col: int, student_end_row: int, lesson_aap_row: int) -> list[SourceExam]:
    attempts = attempt_labels(ws, student_end_row, lesson_aap_row)
    duplicate_counter: Counter[str] = Counter()
    exams: list[SourceExam] = []
    for column in range(3, aap_col):
        label = compact_text(ws.cell(3, column).value) or compact_text(ws.cell(2, column).value)
        if not label:
            continue
        # Skip columns without any student score.
        has_score = any(parse_score(ws.cell(row, column).value) is not None for row in range(4, student_end_row + 1))
        if not has_score:
            continue
        duplicate_counter[normalize_key(label)] += 1
        attempt = attempts.get(column, "")
        if not attempt and duplicate_counter[normalize_key(label)] > 1:
            attempt = f"{duplicate_counter[normalize_key(label)]}"
        exams.append(
            SourceExam(
                column=column,
                order=column,
                label=label,
                topic=label,
                attempt=attempt,
                lesson_order=parse_lesson_order(ws.cell(2, column).value),
                exam_date=parse_source_date(ws.cell(1, column).value),
            )
        )
    return exams


def ensure_lesson_source_columns(conn) -> None:
    conn.execute(
        """
        ALTER TABLE msi_v2.lesson_sessions
            ADD COLUMN IF NOT EXISTS source_key TEXT NOT NULL DEFAULT '',
            ADD COLUMN IF NOT EXISTS source_kind TEXT NOT NULL DEFAULT '',
            ADD COLUMN IF NOT EXISTS source_label TEXT NOT NULL DEFAULT '',
            ADD COLUMN IF NOT EXISTS source_topic TEXT NOT NULL DEFAULT '',
            ADD COLUMN IF NOT EXISTS source_order INTEGER NOT NULL DEFAULT 0,
            ADD COLUMN IF NOT EXISTS source_file TEXT NOT NULL DEFAULT '',
            ADD COLUMN IF NOT EXISTS source_sheet TEXT NOT NULL DEFAULT ''
        """
    )
    conn.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_lesson_sessions_source_key
        ON msi_v2.lesson_sessions (source_key)
        WHERE source_key <> ''
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_lesson_sessions_group_source_order
        ON msi_v2.lesson_sessions (group_id, source_order)
        WHERE source_order > 0
        """
    )


def resolve_group(conn, school_key: str, group_name: str, subject_name: str):
    rows = conn.execute(
        """
        SELECT g.id, g.program_id, s.school_key, g.group_name, subj.subject_name
        FROM msi_v2.groups g
        JOIN msi_v2.schools s ON s.id = g.school_id
        JOIN msi_v2.subject_programs sp ON sp.id = g.program_id
        JOIN msi_v2.subjects subj ON subj.id = sp.subject_id
        WHERE lower(s.school_key) = lower(%s)
          AND lower(g.group_name) = lower(%s)
        """,
        (school_key, group_name),
    ).fetchall()
    subject_key = normalize_key(subject_name)
    for row in rows:
        if normalize_key(row["subject_name"]) == subject_key:
            return row
    return None


def group_students(conn, group_id: int) -> dict[str, dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT gs.group_id, gs.student_id, gs.legacy_enrollment_id, st.full_name
        FROM msi_v2.group_students gs
        JOIN msi_v2.students st ON st.id = gs.student_id
        WHERE gs.group_id = %s AND gs.enrollment_status = 'active'
        """,
        (group_id,),
    ).fetchall()
    return {normalize_key(row["full_name"]): dict(row) for row in rows}


def program_items(conn, program_id: int):
    rows = conn.execute(
        """
        SELECT id, item_order, item_type, lesson_number, title
        FROM msi_v2.subject_program_items
        WHERE program_id = %s
        """,
        (program_id,),
    ).fetchall()
    by_lesson_order: dict[int, dict[str, Any]] = {}
    exams_by_order: dict[int, dict[str, Any]] = {}
    exams_by_title: dict[str, dict[str, Any]] = {}
    for row in rows:
        data = dict(row)
        item_order = int(row["item_order"] or 0)
        if row["item_type"] == "lesson":
            by_lesson_order[item_order] = data
        elif row["item_type"] == "exam":
            exams_by_order[item_order] = data
            exams_by_title[normalize_key(row["title"])] = data
    return by_lesson_order, exams_by_order, exams_by_title


def sync_curriculum_session(conn, *, group_id: int, program_item_id: int, source_key: str, session: SourceSession | SourceExam, source_file: str, source_sheet: str, kind: str, apply: bool):
    existing = conn.execute(
        """
        SELECT id FROM msi_v2.lesson_sessions
        WHERE source_key = %s
           OR (group_id = %s AND program_item_id = %s)
        ORDER BY CASE
                   WHEN source_key = %s THEN 0
                   WHEN source_key <> '' THEN 1
                   ELSE 2
                 END,
                 id
        LIMIT 1
        """,
        (source_key, group_id, program_item_id, source_key),
    ).fetchone()
    status = "completed" if session.session_date else "scheduled"
    if existing:
        if apply:
            conn.execute(
                """
                UPDATE msi_v2.lesson_sessions
                SET group_id = %s,
                    program_item_id = %s,
                    session_date = %s,
                    status = %s,
                    source_key = %s,
                    source_kind = %s,
                    source_label = %s,
                    source_topic = %s,
                    source_order = %s,
                    source_file = %s,
                    source_sheet = %s,
                    updated_at = now()
                WHERE id = %s
                """,
                (
                    group_id,
                    program_item_id,
                    session.session_date,
                    status,
                    source_key,
                    kind,
                    session.label,
                    session.topic,
                    session.order,
                    source_file,
                    source_sheet,
                    int(existing["id"]),
                ),
            )
        return int(existing["id"])
    if not apply:
        return 0
    row = conn.execute(
        """
        INSERT INTO msi_v2.lesson_sessions (
            group_id, program_item_id, session_date, status,
            source_key, source_kind, source_label, source_topic,
            source_order, source_file, source_sheet
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
        """,
        (
            group_id,
            program_item_id,
            session.session_date,
            status,
            source_key,
            kind,
            session.label,
            session.topic,
            session.order,
            source_file,
            source_sheet,
        ),
    ).fetchone()
    return int(row["id"])


def sync_source_session(conn, *, group_id: int, source_key: str, session: SourceSession, source_file: str, source_sheet: str, apply: bool) -> int:
    status = "cancelled" if session.kind == "cancelled" else "completed"
    if not apply:
        row = conn.execute("SELECT id FROM msi_v2.lesson_sessions WHERE source_key = %s", (source_key,)).fetchone()
        return int(row["id"]) if row else 0
    row = conn.execute(
        """
        INSERT INTO msi_v2.lesson_sessions (
            group_id, session_date, status, source_key, source_kind,
            source_label, source_topic, source_order, source_file, source_sheet
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (source_key) WHERE source_key <> '' DO UPDATE SET
            group_id = excluded.group_id,
            session_date = excluded.session_date,
            status = excluded.status,
            source_kind = excluded.source_kind,
            source_label = excluded.source_label,
            source_topic = excluded.source_topic,
            source_order = excluded.source_order,
            source_file = excluded.source_file,
            source_sheet = excluded.source_sheet,
            updated_at = now()
        RETURNING id
        """,
        (
            group_id,
            session.session_date,
            status,
            source_key,
            session.kind,
            session.label,
            session.topic,
            session.order,
            source_file,
            source_sheet,
        ),
    ).fetchone()
    return int(row["id"])


def upsert_attendance(conn, *, lesson_session_id: int, group_id: int, student_id: int, status: str, apply: bool) -> str:
    if not lesson_session_id:
        return "skip"
    if not apply:
        return "upsert" if status else "delete"
    if status:
        conn.execute(
            """
            INSERT INTO msi_v2.attendance_records (
                lesson_session_id, group_id, student_id, attendance_status, created_at, updated_at
            )
            VALUES (%s, %s, %s, %s, now(), now())
            ON CONFLICT (lesson_session_id, student_id) DO UPDATE SET
                attendance_status = excluded.attendance_status,
                updated_at = now()
            """,
            (lesson_session_id, group_id, student_id, status),
        )
        return "upsert"
    conn.execute(
        "DELETE FROM msi_v2.attendance_records WHERE lesson_session_id = %s AND student_id = %s",
        (lesson_session_id, student_id),
    )
    return "delete"


def upsert_homework(conn, *, lesson_session_id: int, group_id: int, student_id: int, score: float | None, apply: bool) -> str:
    if not lesson_session_id:
        return "skip"
    if not apply:
        return "upsert" if score is not None else "delete"
    if score is not None:
        conn.execute(
            """
            INSERT INTO msi_v2.homework_scores (
                lesson_session_id, group_id, student_id, score, score_scale, created_at, updated_at
            )
            VALUES (%s, %s, %s, %s, 9, now(), now())
            ON CONFLICT (lesson_session_id, student_id) DO UPDATE SET
                score = excluded.score,
                updated_at = now()
            """,
            (lesson_session_id, group_id, student_id, score),
        )
        return "upsert"
    conn.execute(
        "DELETE FROM msi_v2.homework_scores WHERE lesson_session_id = %s AND student_id = %s",
        (lesson_session_id, student_id),
    )
    return "delete"


def upsert_exam(conn, *, group_id: int, student_id: int, program_item_id: int | None, exam: SourceExam, score: float | None, apply: bool) -> str:
    if not apply:
        return "upsert" if score is not None else "delete"
    if score is not None:
        conn.execute(
            """
            INSERT INTO msi_v2.exam_results (
                group_id, program_item_id, student_id, exam_name, attempt,
                score, score_scale, created_at, updated_at
            )
            SELECT %s, %s, %s, %s, %s, %s, 9, now(), now()
            WHERE NOT EXISTS (
                SELECT 1 FROM msi_v2.exam_results
                WHERE group_id = %s
                  AND student_id = %s
                  AND lower(exam_name) = lower(%s)
                  AND lower(attempt) = lower(%s)
            )
            """,
            (
                group_id,
                program_item_id,
                student_id,
                exam.label,
                exam.attempt,
                score,
                group_id,
                student_id,
                exam.label,
                exam.attempt,
            ),
        )
        conn.execute(
            """
            UPDATE msi_v2.exam_results
            SET score = %s,
                program_item_id = COALESCE(%s, program_item_id),
                updated_at = now()
            WHERE group_id = %s
              AND student_id = %s
              AND lower(exam_name) = lower(%s)
              AND lower(attempt) = lower(%s)
            """,
            (score, program_item_id, group_id, student_id, exam.label, exam.attempt),
        )
        return "upsert"
    conn.execute(
        """
        DELETE FROM msi_v2.exam_results
        WHERE group_id = %s
          AND student_id = %s
          AND lower(exam_name) = lower(%s)
          AND lower(attempt) = lower(%s)
        """,
        (group_id, student_id, exam.label, exam.attempt),
    )
    return "delete"


def source_key(*parts: Any) -> str:
    return "excel:" + ":".join(re.sub(r"[^a-z0-9]+", "-", normalize_key(part)).strip("-") for part in parts)


def sync_sheet(conn, *, path: Path, school_key: str, sheet_name: str, group_name: str, subject_name: str, apply: bool) -> Counter:
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        group = resolve_group(conn, school_key, group_name, subject_name)
        stats: Counter = Counter()
        if not group:
            stats["groups_missing"] += 1
            print(f"MISS group: {school_key} {group_name} {subject_name} ({sheet_name})")
            return stats

        aap_col = find_aap_column(ws)
        lesson_aap_row = find_lesson_aap_row(ws)
        students = source_students(ws, lesson_aap_row)
        if not students:
            stats["empty_sheets"] += 1
            return stats
        student_end_row = max(student.row for student in students)
        sessions = source_sessions(ws, aap_col)
        exams = source_exams(ws, aap_col, student_end_row, lesson_aap_row)

        enrollments = group_students(conn, int(group["id"]))
        lesson_items, exam_items_by_order, exam_items_by_title = program_items(conn, int(group["program_id"]))

        matched_students: list[tuple[SourceStudent, dict[str, Any]]] = []
        for student in students:
            enrollment = enrollments.get(student.key)
            if not enrollment:
                stats["students_unmatched"] += 1
                print(f"MISS student: {school_key} {sheet_name} -> {student.name}")
                continue
            matched_students.append((student, enrollment))
        stats["students_matched"] += len(matched_students)

        session_ids: dict[int, int] = {}
        for session in sessions:
            key = source_key(school_key, sheet_name, session.column, session.label)
            if session.kind == "lesson" and is_exact_lesson_label(session.label):
                item = lesson_items.get(session.lesson_order)
                if not item:
                    stats["lessons_unmatched"] += 1
                    session_id = sync_source_session(
                        conn,
                        group_id=int(group["id"]),
                        source_key=key,
                        session=session,
                        source_file=path.name,
                        source_sheet=sheet_name,
                        apply=apply,
                    )
                else:
                    session_id = sync_curriculum_session(
                        conn,
                        group_id=int(group["id"]),
                        program_item_id=int(item["id"]),
                        source_key=source_key(school_key, sheet_name, "lesson", session.lesson_order),
                        session=session,
                        source_file=path.name,
                        source_sheet=sheet_name,
                        kind="lesson",
                        apply=apply,
                    )
            else:
                session_id = sync_source_session(
                    conn,
                    group_id=int(group["id"]),
                    source_key=key,
                    session=session,
                    source_file=path.name,
                    source_sheet=sheet_name,
                    apply=apply,
                )
            session_ids[session.column] = session_id
            stats[f"sessions_{session.kind}"] += 1

        for session in sessions:
            if session.kind == "cancelled":
                continue
            session_id = session_ids.get(session.column, 0)
            for student, enrollment in matched_students:
                status = normalize_attendance(ws.cell(student.row, session.attendance_col).value)
                stats[f"attendance_{upsert_attendance(conn, lesson_session_id=session_id, group_id=int(group['id']), student_id=int(enrollment['student_id']), status=status, apply=apply)}"] += 1
                if session.homework_col:
                    score = parse_score(ws.cell(student.row, session.homework_col).value)
                    stats[f"homework_{upsert_homework(conn, lesson_session_id=session_id, group_id=int(group['id']), student_id=int(enrollment['student_id']), score=score, apply=apply)}"] += 1

        for exam in exams:
            item = exam_items_by_order.get(exam.lesson_order) or exam_items_by_title.get(normalize_key(exam.label))
            program_item_id = int(item["id"]) if item else None
            if program_item_id:
                sync_curriculum_session(
                    conn,
                    group_id=int(group["id"]),
                    program_item_id=program_item_id,
                    source_key=source_key(school_key, sheet_name, "exam", exam.lesson_order or exam.label),
                    session=exam,
                    source_file=path.name,
                    source_sheet=sheet_name,
                    kind="exam",
                    apply=apply,
                )
            else:
                stats["exams_unmatched_program_item"] += 1
            for student, enrollment in matched_students:
                score = parse_score(ws.cell(student.row, exam.column).value)
                stats[f"exam_{upsert_exam(conn, group_id=int(group['id']), student_id=int(enrollment['student_id']), program_item_id=program_item_id, exam=exam, score=score, apply=apply)}"] += 1

        print(
            f"{'APPLY' if apply else 'DRY'} {school_key}/{sheet_name}: "
            f"{len(matched_students)}/{len(students)} students, "
            f"{len(sessions)} sessions, {len(exams)} exam columns"
        )
        return stats
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write changes to the database.")
    parser.add_argument("--skip-online", action="store_true", help="Skip School 5 online sheet.")
    args = parser.parse_args()

    totals: Counter = Counter()
    with queries.connect_auth_db() as conn:
        ensure_lesson_source_columns(conn)
        for source in DEFAULT_SOURCES:
            path = source["path"]
            if not path.exists():
                raise FileNotFoundError(path)
            for sheet_name, (group_name, subject_name) in source["sheets"].items():
                if args.skip_online and group_name.casefold() == "online":
                    continue
                totals.update(
                    sync_sheet(
                        conn,
                        path=path,
                        school_key=source["school_key"],
                        sheet_name=sheet_name,
                        group_name=group_name,
                        subject_name=subject_name,
                        apply=args.apply,
                    )
                )
        if args.apply:
            conn.commit()
        else:
            conn.rollback()

    print("\nSummary")
    for key in sorted(totals):
        print(f"{key}: {totals[key]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
